import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import glob
from openpyxl import load_workbook
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
file_pattern = os.path.join(script_dir, 'Timeslot by week', 'W*.xlsx')
files = glob.glob(file_pattern)

print(f"Found {len(files)} files to merge:")
for f in files:
    print(f"  - {f}")

dfs = []
for file in files:
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    df = pd.DataFrame(data[1:], columns=data[0])
    
    if len(df.columns) == 31:
        df.insert(7, '0h00', 0)
    if len(df.columns) == 32:
        df = df.iloc[:, :-1]
    
    df.columns = [f'col_{i}' for i in range(len(df.columns))]
    
    dfs.append(df)

merged_df = pd.concat(dfs, ignore_index=True)

print(f"\nTotal rows after merging: {len(merged_df)}")

merged_df['Date_parsed'] = pd.to_datetime(merged_df.iloc[:, 1], errors='coerce', dayfirst=True)

mask = (merged_df['Date_parsed'].dt.year == 2025) & \
       (merged_df['Date_parsed'].dt.day == 11) & \
       (merged_df['Date_parsed'].dt.month <= 10)

if mask.sum() > 0:
    print(f"\nDetected {mask.sum()} rows with potentially misinterpreted dates (YYYY-M-11 pattern)")
    print("Correcting: YYYY-M-11 -> YYYY-11-M")
    
    corrected_dates = merged_df.loc[mask, 'Date_parsed'].apply(
        lambda x: pd.Timestamp(year=x.year, month=11, day=x.month)
    )
    merged_df.loc[mask, 'Date_parsed'] = corrected_dates

nov_data = merged_df[(merged_df['Date_parsed'].dt.year == 2025) & 
                     (merged_df['Date_parsed'].dt.month == 11)]

print(f"Rows with November 2025 data: {len(nov_data)}")

nov_data = nov_data[nov_data.iloc[:, 3].notna()]

print(f"Rows after removing blank Category (column D): {len(nov_data)}")

nov_data = nov_data[nov_data.iloc[:, 5].isin(['Booking taken', 'Available Capacity'])]

print(f"Rows after filtering Booking taken and Available Capacity: {len(nov_data)}")

print("\n=== Saving filtered data to Excel for inspection ===")
nov_data.to_excel('Timeslot_Filtered_Data_November_2025.xlsx', index=False)
print("Filtered data saved to Timeslot_Filtered_Data_November_2025.xlsx")

time_columns = list(range(7, 31))

def extract_timeslot_data(df, condition_type, category):
    if condition_type == 'outbound':
        condition_filter = df.iloc[:, 0] == 'Loading'
    else:
        condition_filter = df.iloc[:, 0] == 'Reception'
    
    category_filter = df.iloc[:, 3] == category
    
    filtered_df = df[condition_filter & category_filter]
    
    booking_taken = filtered_df[filtered_df.iloc[:, 5] == 'Booking taken'].iloc[:, time_columns].sum()
    available_capacity = filtered_df[filtered_df.iloc[:, 5] == 'Available Capacity'].iloc[:, time_columns].sum()
    
    result = pd.DataFrame({
        'Time Slot': range(24),
        'Booking Taken': booking_taken.values,
        'Available Capacity': available_capacity.values
    })
    
    total_capacity = result['Booking Taken'] + result['Available Capacity']
    result['Utilization Rate'] = result['Booking Taken'] / total_capacity.replace(0, np.nan)
    
    return result

def plot_booking_taken(data, category, direction):
    fig, ax = plt.subplots(figsize=(12, 6))
    
    x = data['Time Slot']
    
    color = 'steelblue'
    ax.bar(x, data['Booking Taken'], color=color, alpha=0.7, label='Booking Taken')
    ax.set_xlabel('Time Slot (Hour)', fontsize=12)
    ax.set_ylabel('Booking Taken', fontsize=12)
    ax.set_xticks(range(24))
    ax.set_xticklabels(range(24))
    ax.grid(True, alpha=0.3, axis='y')
    ax.legend(loc='upper left')
    
    plt.title(f'{category} - {direction.capitalize()} - November 2025 Hourly Booking Taken', 
              fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(f'{category}_{direction}_Booking_Taken_November_2025.png', dpi=300, bbox_inches='tight')
    plt.show()

def plot_utilization_stacked(data, category, direction):
    fig, ax = plt.subplots(figsize=(12, 6))
    
    x = data['Time Slot']
    
    booking_taken = data['Booking Taken']
    available = data['Available Capacity']
    
    ax.bar(x, booking_taken, label='Booking Taken', color='steelblue', alpha=0.8)
    ax.bar(x, available, bottom=booking_taken, label='Available Capacity', 
           color='lightcoral', alpha=0.8)
    
    ax.set_xlabel('Time Slot (Hour)', fontsize=12)
    ax.set_ylabel('Capacity', fontsize=12)
    ax.set_xticks(range(24))
    ax.set_xticklabels(range(24))
    ax.legend(loc='upper left')
    ax.grid(True, alpha=0.3, axis='y')
    
    for i, (bt, total) in enumerate(zip(booking_taken, booking_taken + available)):
        if total > 0:
            util_rate = bt / total
            if util_rate > 0.05:
                ax.text(i, total/2, f'{util_rate:.1%}', ha='center', va='center', 
                       fontsize=8, fontweight='bold', color='white')
    
    plt.title(f'{category} - {direction.capitalize()} - November 2025 Slot Utilization', 
              fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(f'{category}_{direction}_Utilization_November_2025.png', dpi=300, bbox_inches='tight')
    plt.show()

print("\n=== Extracting FG Inbound data ===")
fg_inbound = extract_timeslot_data(nov_data, 'inbound', 'FG')
print(fg_inbound.head())

print("\n=== Extracting FG Outbound data ===")
fg_outbound = extract_timeslot_data(nov_data, 'outbound', 'FG')
print(fg_outbound.head())

print("\n=== Extracting R&P Inbound data ===")
rp_inbound = extract_timeslot_data(nov_data, 'inbound', 'R&P')
print(rp_inbound.head())

print("\n=== Extracting R&P Outbound data ===")
rp_outbound = extract_timeslot_data(nov_data, 'outbound', 'R&P')
print(rp_outbound.head())

print("\n=== Plotting Booking Taken charts ===")
plot_booking_taken(fg_inbound, 'FG', 'inbound')
plot_booking_taken(fg_outbound, 'FG', 'outbound')
plot_booking_taken(rp_inbound, 'R&P', 'inbound')
plot_booking_taken(rp_outbound, 'R&P', 'outbound')

print("\n=== Plotting Utilization Rate charts ===")
plot_utilization_stacked(fg_inbound, 'FG', 'inbound')
plot_utilization_stacked(fg_outbound, 'FG', 'outbound')
plot_utilization_stacked(rp_inbound, 'R&P', 'inbound')
plot_utilization_stacked(rp_outbound, 'R&P', 'outbound')

print("\n=== Saving data to Excel ===")
with pd.ExcelWriter('Timeslot_Analysis_Data_November_2025.xlsx', engine='openpyxl') as writer:
    fg_inbound.to_excel(writer, sheet_name='FG_Inbound', index=False)
    fg_outbound.to_excel(writer, sheet_name='FG_Outbound', index=False)
    rp_inbound.to_excel(writer, sheet_name='R&P_Inbound', index=False)
    rp_outbound.to_excel(writer, sheet_name='R&P_Outbound', index=False)

print("\nAll charts and data have been saved!")
