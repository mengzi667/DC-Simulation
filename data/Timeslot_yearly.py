import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
import glob
from openpyxl import load_workbook
import os
import warnings
warnings.filterwarnings('ignore')

# Set default font
plt.rcParams['font.sans-serif'] = ['Arial']
plt.rcParams['axes.unicode_minus'] = False

script_dir = os.path.dirname(os.path.abspath(__file__))
data_folder = os.path.join(script_dir, 'Timeslot_by_week', 'Timeslot by week')
file_pattern = os.path.join(data_folder, 'W*.xlsx')
files = glob.glob(file_pattern)

print(f"Found {len(files)} files to merge:")
for f in sorted(files):
    print(f"  - {os.path.basename(f)}")

# Read and merge all files
dfs = []
for file in files:
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        df = pd.DataFrame(data[1:], columns=data[0])
        
        # Standardize column count
        if len(df.columns) == 31:
            df.insert(7, '0h00', 0)
        if len(df.columns) == 32:
            df = df.iloc[:, :-1]
        
        df.columns = [f'col_{i}' for i in range(len(df.columns))]
        dfs.append(df)
        
    except Exception as e:
        print(f"Error reading {file}: {e}")

merged_df = pd.concat(dfs, ignore_index=True)
print(f"\nTotal rows after merging: {len(merged_df)}")

# Parse dates
merged_df['Date_parsed'] = pd.to_datetime(merged_df.iloc[:, 1], errors='coerce', dayfirst=True)

# Correct date errors
mask = (merged_df['Date_parsed'].dt.year == 2025) & \
       (merged_df['Date_parsed'].dt.day == 11) & \
       (merged_df['Date_parsed'].dt.month <= 10)

if mask.sum() > 0:
    print(f"\nDetected {mask.sum()} rows with potentially misinterpreted dates")
    print("Correcting: YYYY-M-11 -> YYYY-11-M")
    corrected_dates = merged_df.loc[mask, 'Date_parsed'].apply(
        lambda x: pd.Timestamp(year=x.year, month=11, day=x.month)
    )
    merged_df.loc[mask, 'Date_parsed'] = corrected_dates

# Filter 2025 yearly data
year_data = merged_df[merged_df['Date_parsed'].dt.year == 2025]
print(f"Rows with 2025 data: {len(year_data)}")

# Remove blank category rows
year_data = year_data[year_data.iloc[:, 3].notna()]
print(f"Rows after removing blank Category: {len(year_data)}")

# Keep only Booking taken and Available Capacity
year_data = year_data[year_data.iloc[:, 5].isin(['Booking taken', 'Available Capacity'])]
print(f"Rows after filtering Booking taken and Available Capacity: {len(year_data)}")

# Create output directory
output_dir = os.path.join(script_dir, 'Timeslot_Yearly_Analysis')
os.makedirs(output_dir, exist_ok=True)

# Time column indices
time_columns = list(range(7, 31))

def extract_timeslot_data(df, condition_type, category):
    """Extract timeslot data"""
    if condition_type == 'outbound':
        condition_filter = df.iloc[:, 0] == 'Loading'
    else:
        condition_filter = df.iloc[:, 0] == 'Reception'
    
    category_filter = df.iloc[:, 3] == category
    filtered_df = df[condition_filter & category_filter]
    
    booking_taken = filtered_df[filtered_df.iloc[:, 5] == 'Booking taken'].iloc[:, time_columns].sum()
    available_capacity = filtered_df[filtered_df.iloc[:, 5] == 'Available Capacity'].iloc[:, time_columns].sum()
    
    result = pd.DataFrame({
        'Time Slot': range(24),
        'Booking Taken': booking_taken.values,
        'Available Capacity': available_capacity.values
    })
    
    total_capacity = result['Booking Taken'] + result['Available Capacity']
    result['Utilization Rate'] = result['Booking Taken'] / total_capacity.replace(0, np.nan)
    
    return result

def plot_booking_taken(data, category, direction, month_name):
    """Plot booking taken bar chart"""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    x = data['Time Slot']
    color = 'steelblue'
    ax.bar(x, data['Booking Taken'], color=color, alpha=0.7, label='Booking Taken')
    ax.set_xlabel('Time Slot (Hour)', fontsize=12)
    ax.set_ylabel('Booking Taken', fontsize=12)
    ax.set_xticks(range(24))
    ax.set_xticklabels(range(24))
    ax.grid(True, alpha=0.3, axis='y')
    ax.legend(loc='upper left')
    
    plt.title(f'{category} - {direction.capitalize()} - {month_name} 2025 Hourly Booking Taken', 
              fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    filename = f'{category}_{direction}_Booking_Taken_{month_name}_2025.png'
    plt.savefig(os.path.join(output_dir, filename), dpi=300, bbox_inches='tight')
    plt.close()

def plot_utilization_stacked(data, category, direction, month_name):
    """Plot utilization rate stacked chart"""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    x = data['Time Slot']
    booking_taken = data['Booking Taken']
    available = data['Available Capacity']
    
    ax.bar(x, booking_taken, label='Booking Taken', color='steelblue', alpha=0.8)
    ax.bar(x, available, bottom=booking_taken, label='Available Capacity', 
           color='lightcoral', alpha=0.8)
    
    ax.set_xlabel('Time Slot (Hour)', fontsize=12)
    ax.set_ylabel('Capacity', fontsize=12)
    ax.set_xticks(range(24))
    ax.set_xticklabels(range(24))
    ax.legend(loc='upper left')
    ax.grid(True, alpha=0.3, axis='y')
    
    # Add utilization rate labels
    for i, (bt, total) in enumerate(zip(booking_taken, booking_taken + available)):
        if total > 0:
            util_rate = bt / total
            if util_rate > 0.05:
                ax.text(i, total/2, f'{util_rate:.0%}', ha='center', va='center', 
                       fontsize=7, fontweight='bold', color='white')
    
    plt.title(f'{category} - {direction.capitalize()} - {month_name} 2025 Slot Utilization', 
              fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    filename = f'{category}_{direction}_Utilization_{month_name}_2025.png'
    plt.savefig(os.path.join(output_dir, filename), dpi=300, bbox_inches='tight')
    plt.close()

# Month name mapping
month_names = {
    1: 'January', 2: 'February', 3: 'March', 4: 'April',
    5: 'May', 6: 'June', 7: 'July', 8: 'August',
    9: 'September', 10: 'October', 11: 'November', 12: 'December'
}

# Store yearly data for boxplots
yearly_utilization_data = {
    'FG_inbound': [],
    'FG_outbound': [],
    'R&P_inbound': [],
    'R&P_outbound': []
}

yearly_booking_data = {
    'FG_inbound': [],
    'FG_outbound': [],
    'R&P_inbound': [],
    'R&P_outbound': []
}

yearly_available_data = {
    'FG_inbound': [],
    'FG_outbound': [],
    'R&P_inbound': [],
    'R&P_outbound': []
}

# Generate charts for each month
print("\n=== Processing monthly data ===")
for month_num in range(1, 13):
    month_name = month_names[month_num]
    print(f"\nProcessing {month_name}...")
    
    # Filter data for this month
    month_data = year_data[year_data['Date_parsed'].dt.month == month_num]
    
    if len(month_data) == 0:
        print(f"  No data found for {month_name}, skipping...")
        continue
    
    print(f"  Rows: {len(month_data)}")
    
    # Extract data for each category
    fg_inbound = extract_timeslot_data(month_data, 'inbound', 'FG')
    fg_outbound = extract_timeslot_data(month_data, 'outbound', 'FG')
    rp_inbound = extract_timeslot_data(month_data, 'inbound', 'R&P')
    rp_outbound = extract_timeslot_data(month_data, 'outbound', 'R&P')
    
    # Collect utilization data for boxplots
    yearly_utilization_data['FG_inbound'].append({
        'month': month_name,
        'rates': fg_inbound['Utilization Rate'].dropna().values
    })
    yearly_utilization_data['FG_outbound'].append({
        'month': month_name,
        'rates': fg_outbound['Utilization Rate'].dropna().values
    })
    yearly_utilization_data['R&P_inbound'].append({
        'month': month_name,
        'rates': rp_inbound['Utilization Rate'].dropna().values
    })
    yearly_utilization_data['R&P_outbound'].append({
        'month': month_name,
        'rates': rp_outbound['Utilization Rate'].dropna().values
    })
    
    # Collect booking taken data for boxplots
    yearly_booking_data['FG_inbound'].append({
        'month': month_name,
        'values': fg_inbound['Booking Taken'].values
    })
    yearly_booking_data['FG_outbound'].append({
        'month': month_name,
        'values': fg_outbound['Booking Taken'].values
    })
    yearly_booking_data['R&P_inbound'].append({
        'month': month_name,
        'values': rp_inbound['Booking Taken'].values
    })
    yearly_booking_data['R&P_outbound'].append({
        'month': month_name,
        'values': rp_outbound['Booking Taken'].values
    })
    
    # Collect available capacity data for boxplots
    yearly_available_data['FG_inbound'].append({
        'month': month_name,
        'values': fg_inbound['Available Capacity'].values
    })
    yearly_available_data['FG_outbound'].append({
        'month': month_name,
        'values': fg_outbound['Available Capacity'].values
    })
    yearly_available_data['R&P_inbound'].append({
        'month': month_name,
        'values': rp_inbound['Available Capacity'].values
    })
    yearly_available_data['R&P_outbound'].append({
        'month': month_name,
        'values': rp_outbound['Available Capacity'].values
    })
    
    # Generate charts
    plot_booking_taken(fg_inbound, 'FG', 'inbound', month_name)
    plot_booking_taken(fg_outbound, 'FG', 'outbound', month_name)
    plot_booking_taken(rp_inbound, 'R&P', 'inbound', month_name)
    plot_booking_taken(rp_outbound, 'R&P', 'outbound', month_name)
    
    plot_utilization_stacked(fg_inbound, 'FG', 'inbound', month_name)
    plot_utilization_stacked(fg_outbound, 'FG', 'outbound', month_name)
    plot_utilization_stacked(rp_inbound, 'R&P', 'inbound', month_name)
    plot_utilization_stacked(rp_outbound, 'R&P', 'outbound', month_name)
    
    print(f"  Generated 8 charts for {month_name}")

# Generate yearly boxplots
print("\n=== Generating yearly boxplots ===")

def plot_yearly_boxplot(util_data, category, direction):
    """Plot yearly utilization rate boxplot"""
    fig, ax = plt.subplots(figsize=(14, 7))
    
    # Prepare data
    data_to_plot = []
    labels = []
    
    for month_data in util_data:
        if len(month_data['rates']) > 0:
            data_to_plot.append(month_data['rates'])
            labels.append(month_data['month'][:3])  # Use month abbreviation
    
    if len(data_to_plot) == 0:
        print(f"  No data for {category} {direction}")
        plt.close()
        return
    
    # Draw boxplot
    bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True,
                     showmeans=True, meanline=True)
    
    # Customize boxplot appearance
    for patch in bp['boxes']:
        patch.set_facecolor('lightblue')
        patch.set_alpha(0.7)
    
    for whisker in bp['whiskers']:
        whisker.set(color='gray', linewidth=1.5, linestyle=':')
    
    for cap in bp['caps']:
        cap.set(color='gray', linewidth=1.5)
    
    for median in bp['medians']:
        median.set(color='red', linewidth=2)
    
    for mean in bp['means']:
        mean.set(color='green', linewidth=2)
    
    ax.set_xlabel('Month', fontsize=12)
    ax.set_ylabel('Utilization Rate', fontsize=12)
    ax.set_title(f'{category} - {direction.capitalize()} - 2025 Yearly Utilization Rate Distribution', 
                 fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, axis='y')
    ax.set_ylim(0, 1)
    
    # Add legend
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color='red', linewidth=2, label='Median'),
        Line2D([0], [0], color='green', linewidth=2, label='Mean')
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    
    plt.tight_layout()
    filename = f'{category}_{direction}_Yearly_Boxplot_2025.png'
    plt.savefig(os.path.join(output_dir, filename), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Generated {filename}")

# Generate boxplots for each category and direction
plot_yearly_boxplot(yearly_utilization_data['FG_inbound'], 'FG', 'inbound')
plot_yearly_boxplot(yearly_utilization_data['FG_outbound'], 'FG', 'outbound')
plot_yearly_boxplot(yearly_utilization_data['R&P_inbound'], 'R&P', 'inbound')
plot_yearly_boxplot(yearly_utilization_data['R&P_outbound'], 'R&P', 'outbound')

print("\n=== Generating yearly boxplots for Booking Taken ===")

def plot_yearly_boxplot_booking(booking_data, category, direction):
    """Plot yearly booking taken boxplot"""
    fig, ax = plt.subplots(figsize=(14, 7))
    
    # Prepare data
    data_to_plot = []
    labels = []
    
    for month_data in booking_data:
        if len(month_data['values']) > 0:
            data_to_plot.append(month_data['values'])
            labels.append(month_data['month'][:3])  # Use month abbreviation
    
    if len(data_to_plot) == 0:
        print(f"  No data for {category} {direction}")
        plt.close()
        return
    
    # Draw boxplot
    bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True,
                     showmeans=True, meanline=True)
    
    # Customize boxplot appearance
    for patch in bp['boxes']:
        patch.set_facecolor('lightgreen')
        patch.set_alpha(0.7)
    
    for whisker in bp['whiskers']:
        whisker.set(color='gray', linewidth=1.5, linestyle=':')
    
    for cap in bp['caps']:
        cap.set(color='gray', linewidth=1.5)
    
    for median in bp['medians']:
        median.set(color='red', linewidth=2)
    
    for mean in bp['means']:
        mean.set(color='blue', linewidth=2)
    
    ax.set_xlabel('Month', fontsize=12)
    ax.set_ylabel('Booking Taken', fontsize=12)
    ax.set_title(f'{category} - {direction.capitalize()} - 2025 Yearly Booking Taken Distribution', 
                 fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, axis='y')
    
    # Add legend
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color='red', linewidth=2, label='Median'),
        Line2D([0], [0], color='blue', linewidth=2, label='Mean')
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    
    plt.tight_layout()
    filename = f'{category}_{direction}_Booking_Taken_Yearly_Boxplot_2025.png'
    plt.savefig(os.path.join(output_dir, filename), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Generated {filename}")

# Generate booking taken boxplots
plot_yearly_boxplot_booking(yearly_booking_data['FG_inbound'], 'FG', 'inbound')
plot_yearly_boxplot_booking(yearly_booking_data['FG_outbound'], 'FG', 'outbound')
plot_yearly_boxplot_booking(yearly_booking_data['R&P_inbound'], 'R&P', 'inbound')
plot_yearly_boxplot_booking(yearly_booking_data['R&P_outbound'], 'R&P', 'outbound')

print("\n=== Generating yearly boxplots for Available Capacity ===")

def plot_yearly_boxplot_available(available_data, category, direction):
    """Plot yearly available capacity boxplot"""
    fig, ax = plt.subplots(figsize=(14, 7))
    
    # Prepare data
    data_to_plot = []
    labels = []
    
    for month_data in available_data:
        if len(month_data['values']) > 0:
            data_to_plot.append(month_data['values'])
            labels.append(month_data['month'][:3])  # Use month abbreviation
    
    if len(data_to_plot) == 0:
        print(f"  No data for {category} {direction}")
        plt.close()
        return
    
    # Draw boxplot
    bp = ax.boxplot(data_to_plot, labels=labels, patch_artist=True,
                     showmeans=True, meanline=True)
    
    # Customize boxplot appearance
    for patch in bp['boxes']:
        patch.set_facecolor('lightcoral')
        patch.set_alpha(0.7)
    
    for whisker in bp['whiskers']:
        whisker.set(color='gray', linewidth=1.5, linestyle=':')
    
    for cap in bp['caps']:
        cap.set(color='gray', linewidth=1.5)
    
    for median in bp['medians']:
        median.set(color='red', linewidth=2)
    
    for mean in bp['means']:
        mean.set(color='blue', linewidth=2)
    
    ax.set_xlabel('Month', fontsize=12)
    ax.set_ylabel('Available Capacity', fontsize=12)
    ax.set_title(f'{category} - {direction.capitalize()} - 2025 Yearly Available Capacity Distribution', 
                 fontsize=14, fontweight='bold')
    ax.grid(True, alpha=0.3, axis='y')
    
    # Add legend
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], color='red', linewidth=2, label='Median'),
        Line2D([0], [0], color='blue', linewidth=2, label='Mean')
    ]
    ax.legend(handles=legend_elements, loc='upper right')
    
    plt.tight_layout()
    filename = f'{category}_{direction}_Available_Capacity_Yearly_Boxplot_2025.png'
    plt.savefig(os.path.join(output_dir, filename), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"  Generated {filename}")

# Generate available capacity boxplots
plot_yearly_boxplot_available(yearly_available_data['FG_inbound'], 'FG', 'inbound')
plot_yearly_boxplot_available(yearly_available_data['FG_outbound'], 'FG', 'outbound')
plot_yearly_boxplot_available(yearly_available_data['R&P_inbound'], 'R&P', 'inbound')
plot_yearly_boxplot_available(yearly_available_data['R&P_outbound'], 'R&P', 'outbound')

# Save summary data to Excel
print("\n=== Saving summary data to Excel ===")
summary_data = []

for month_num in range(1, 13):
    month_name = month_names[month_num]
    month_data = year_data[year_data['Date_parsed'].dt.month == month_num]
    
    if len(month_data) == 0:
        continue
    
    fg_inbound = extract_timeslot_data(month_data, 'inbound', 'FG')
    fg_outbound = extract_timeslot_data(month_data, 'outbound', 'FG')
    rp_inbound = extract_timeslot_data(month_data, 'inbound', 'R&P')
    rp_outbound = extract_timeslot_data(month_data, 'outbound', 'R&P')
    
    summary_data.append({
        'Month': month_name,
        'FG_Inbound_Avg_Utilization': fg_inbound['Utilization Rate'].mean(),
        'FG_Outbound_Avg_Utilization': fg_outbound['Utilization Rate'].mean(),
        'R&P_Inbound_Avg_Utilization': rp_inbound['Utilization Rate'].mean(),
        'R&P_Outbound_Avg_Utilization': rp_outbound['Utilization Rate'].mean(),
        'FG_Inbound_Total_Bookings': fg_inbound['Booking Taken'].sum(),
        'FG_Outbound_Total_Bookings': fg_outbound['Booking Taken'].sum(),
        'R&P_Inbound_Total_Bookings': rp_inbound['Booking Taken'].sum(),
        'R&P_Outbound_Total_Bookings': rp_outbound['Booking Taken'].sum(),
    })

summary_df = pd.DataFrame(summary_data)
summary_file = os.path.join(output_dir, 'Yearly_Summary_2025.xlsx')
summary_df.to_excel(summary_file, index=False)
print(f"Summary data saved to {summary_file}")

print("\n" + "="*60)
print("All analysis complete!")
print(f"Total charts generated: {len(glob.glob(os.path.join(output_dir, '*.png')))}")
print(f"Output directory: {output_dir}")
print("="*60)
