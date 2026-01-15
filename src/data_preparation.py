"""DC仿真数据准备脚本"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
import json
import os

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / 'data' / 'raw'
TIMESLOT_DIR = PROJECT_ROOT / 'data' / 'Timeslot by week'
OUTPUT_DIR = PROJECT_ROOT / 'outputs' / 'simulation_configs'
FIGURES_DIR = PROJECT_ROOT / 'outputs' / 'figures'

KPI_FILE = DATA_DIR / 'KPI sheet 2025.xlsx'
SHIPMENTS_FILE = DATA_DIR / 'Total Shipments 2025.xlsx'

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
FIGURES_DIR.mkdir(parents=True, exist_ok=True)

FTE_DATA_FILE = DATA_DIR / 'Input FTE Data.txt'

def extract_fte_from_file():
    """从Input FTE Data.txt提取FTE配置"""
    print("\n" + "=" * 60)
    print("提取FTE配置")
    print("=" * 60)
    
    if not FTE_DATA_FILE.exists():
        print(f"警告: FTE数据文件不存在: {FTE_DATA_FILE}")
        return None
    
    try:
        with open(FTE_DATA_FILE, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 解析FTE数据
        fte_config = {
            'FG': {},
            'R&P': {}
        }
        
        lines = content.split('\n')
        current_category = None
        
        for line in lines:
            line = line.strip()
            
            # 检查是否是FG行（FG: Inbound FTE: ...）
            if line.startswith('FG:'):
                current_category = 'FG'
                # 直接从这行提取Inbound FTE
                if 'Inbound FTE:' in line:
                    parts = line.split('Inbound FTE:')
                    if len(parts) > 1:
                        value = float(parts[1].strip())
                        fte_config['FG']['Inbound'] = value
            elif line.startswith('R&P:'):
                current_category = 'R&P'
            elif 'Inbound FTE:' in line and current_category:
                # 处理缩进的Inbound FTE行
                value = float(line.split(':')[-1].strip())
                fte_config[current_category]['Inbound'] = value
            elif 'Outbound FTE:' in line and current_category:
                value = float(line.split(':')[-1].strip())
                fte_config[current_category]['Outbound'] = value
            elif 'Before ready to load work efficiency:' in line and current_category:
                parts = line.split(':')[-1].strip().split()
                value = float(parts[0])
                fte_config[current_category]['efficiency'] = value
        
        print(f"\n已读取FTE配置:")
        print(f"  FG:")
        print(f"    Inbound FTE: {fte_config['FG']['Inbound']}")
        print(f"    Outbound FTE: {fte_config['FG']['Outbound']}")
        print(f"    效率: {fte_config['FG']['efficiency']} pallet/FTE")
        print(f"\n  R&P:")
        print(f"    Inbound FTE: {fte_config['R&P']['Inbound']}")
        print(f"    Outbound FTE: {fte_config['R&P']['Outbound']}")
        print(f"    效率: {fte_config['R&P']['efficiency']} pallet/FTE")
        
        # 计算月度工时（22天 × 8小时）
        hours_per_month = 176
        
        # 计算每小时处理能力
        print(f"\n每小时处理能力（基准18小时运营）:")
        for category in ['FG', 'R&P']:
            for direction in ['Inbound', 'Outbound']:
                fte = fte_config[category][direction]
                efficiency = fte_config[category]['efficiency']
                hourly_capacity = (fte * efficiency) / hours_per_month
                print(f"  {category} {direction}: {hourly_capacity:.1f} pallet/h")
        
        return fte_config
        
    except Exception as e:
        print(f"错误: 无法解析FTE数据文件 - {e}")
        import traceback
        traceback.print_exc()
        return None


def extract_efficiency_parameters():
    """提取效率参数（托盘/小时）"""
    print("=" * 60)
    print("1. 提取效率参数")
    print("=" * 60)
    
    df = pd.read_excel(KPI_FILE, sheet_name='Hours & volumes per subgroup', header=None)
    
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    def extract_category_efficiency(df, category):
        hours_data = None
        inbound_data = None
        outbound_data = None
        
        if category == 'R&P':
            for idx, row in df.iterrows():
                if row[0] == 'R&P' and row[1] == 'Actual':
                    hours_data = row[2:14].values
                    break
            
            for idx, row in df.iterrows():
                if row[0] == 'R&P - pallets - inbound' and row[1] == 'Actual':
                    inbound_data = row[2:14].values
                elif row[0] == 'R&P - pallets - outbound' and row[1] == 'Actual':
                    outbound_data = row[2:14].values
                    
        elif category == 'FG':
            for idx, row in df.iterrows():
                if row[0] == 'FG' and row[1] == 'Actual':
                    hours_data = row[2:14].values
                    break
            
            for idx, row in df.iterrows():
                if row[0] == 'FG - pallets - inbound' and row[1] == 'Actual':
                    inbound_data = row[2:14].values
                elif row[0] == 'FG - pallets - outbound' and row[1] == 'Actual':
                    outbound_data = row[2:14].values
        
        hours_data = pd.to_numeric(hours_data, errors='coerce')
        inbound_data = pd.to_numeric(inbound_data, errors='coerce')
        outbound_data = pd.to_numeric(outbound_data, errors='coerce')
        
        total_pallets = inbound_data + outbound_data
        efficiency = total_pallets / hours_data
        valid_efficiency = efficiency[~np.isnan(efficiency)]
        
        result = {
            'mean': np.mean(valid_efficiency),
            'std': np.std(valid_efficiency),
            'median': np.median(valid_efficiency),
            'min': np.min(valid_efficiency),
            'max': np.max(valid_efficiency),
            'monthly_data': list(valid_efficiency)
        }
        
        return result
    
    # 提取 R&P 和 FG 数据
    rp_efficiency = extract_category_efficiency(df, 'R&P')
    fg_efficiency = extract_category_efficiency(df, 'FG')
    
    efficiency_params = {
        'R&P': rp_efficiency,
        'FG': fg_efficiency
    }
    
    print(f"\nR&P 效率参数:")
    print(f"  平均值: {rp_efficiency['mean']:.2f} 托盘/小时")
    print(f"  标准差: {rp_efficiency['std']:.3f}")
    print(f"  中位数: {rp_efficiency['median']:.2f}")
    print(f"  范围: [{rp_efficiency['min']:.2f}, {rp_efficiency['max']:.2f}]")
    
    print(f"\nFG 效率参数:")
    print(f"  平均值: {fg_efficiency['mean']:.2f} 托盘/小时")
    print(f"  标准差: {fg_efficiency['std']:.3f}")
    print(f"  中位数: {fg_efficiency['median']:.2f}")
    print(f"  范围: [{fg_efficiency['min']:.2f}, {fg_efficiency['max']:.2f}]")
    
    return efficiency_params


def extract_demand_distribution(target_year=2025):
    """
    从 Total Shipments 提取需求分布（全年数据）
    
    Args:
        target_year: 目标年份（默认 2025）
    
    Returns:
        dict: 包含每小时到达率和每日需求分布
    """
    print("\n" + "=" * 60)
    print("2. 提取需求分布参数（全年数据）")
    print("=" * 60)
    
    # 读取数据
    inbound_df = pd.read_excel(SHIPMENTS_FILE, sheet_name='Inbound Shipments 2025')
    outbound_df = pd.read_excel(SHIPMENTS_FILE, sheet_name='Outbound Shipments 2025')
    
    # 转换日期
    inbound_df['Date Hour appointement'] = pd.to_datetime(inbound_df['Date Hour appointement'])
    outbound_df['Date Hour appointement'] = pd.to_datetime(outbound_df['Date Hour appointement'])
    
    # 使用全年数据
    inbound_year = inbound_df[inbound_df['Date Hour appointement'].dt.year == target_year]
    outbound_year = outbound_df[outbound_df['Date Hour appointement'].dt.year == target_year]
    
    print(f"\n{target_year} 年全年数据统计:")
    print(f"  入库记录数: {len(inbound_year)}")
    print(f"  出库记录数: {len(outbound_year)}")
    
    # 计算数据覆盖的月份
    inbound_months = inbound_year['Date Hour appointement'].dt.month.unique()
    outbound_months = outbound_year['Date Hour appointement'].dt.month.unique()
    print(f"  入库数据覆盖月份: {sorted(inbound_months.tolist())}")
    print(f"  出库数据覆盖月份: {sorted(outbound_months.tolist())}")
    
    # 提取小时到达分布
    outbound_year['Hour'] = outbound_year['Date Hour appointement'].dt.hour
    inbound_year['Hour'] = inbound_year['Date Hour appointement'].dt.hour
    
    # 分类别统计 Outbound
    hourly_arrivals_outbound = {}
    
    for category in ['FG', 'R&P']:
        category_data = outbound_year[outbound_year['Category'] == category]
        
        # 计算每小时平均到达卡车数（全年平均）
        hourly_counts = category_data.groupby('Hour').size()
        num_days = category_data['Date Hour appointement'].dt.date.nunique()
        
        hourly_rate = (hourly_counts / num_days).to_dict()
        hourly_arrivals_outbound[category] = hourly_rate
        
        total_trucks = hourly_counts.sum()
        avg_per_day = total_trucks / num_days
        print(f"\n{category} 出库统计（全年）:")
        print(f"  总卡车数: {total_trucks}")
        print(f"  统计天数: {num_days}")
        print(f"  平均每天: {avg_per_day:.2f} 辆")
        print(f"  每小时平均到达卡车数:")
        for hour in sorted(hourly_rate.keys()):
            print(f"    {hour:02d}:00 - {hourly_rate[hour]:.2f}")
    
    # 分类别统计 Inbound
    hourly_arrivals_inbound = {}
    
    for category in ['FG', 'R&P']:
        category_data = inbound_year[inbound_year['Category'] == category]
        
        # 计算每小时平均到达卡车数（全年平均）
        hourly_counts = category_data.groupby('Hour').size()
        num_days = category_data['Date Hour appointement'].dt.date.nunique()
        
        hourly_rate = (hourly_counts / num_days).to_dict()
        hourly_arrivals_inbound[category] = hourly_rate
        
        total_trucks = hourly_counts.sum()
        avg_per_day = total_trucks / num_days
        print(f"\n{category} 入库统计（全年）:")
        print(f"  总卡车数: {total_trucks}")
        print(f"  统计天数: {num_days}")
        print(f"  平均每天: {avg_per_day:.2f} 辆")
        print(f"  每小时平均到达卡车数:")
        for hour in sorted(hourly_rate.keys()):
            print(f"    {hour:02d}:00 - {hourly_rate[hour]:.2f}")
    
    # 合并所有类别计算总到达率
    combined_hourly = outbound_year.groupby('Hour').size()
    num_days_total = outbound_year['Date Hour appointement'].dt.date.nunique()
    total_hourly_rate = (combined_hourly / num_days_total).to_dict()
    
    print(f"\n总体每小时平均到达卡车数:")
    for hour in sorted(total_hourly_rate.keys()):
        print(f"  {hour:02d}:00 - {total_hourly_rate[hour]:.2f}")
    
    def daily_demand_stats(df, category):
        cat_data = df[df['Category'] == category].copy()
        cat_data['Date'] = cat_data['Date Hour appointement'].dt.date
        
        daily_pallets = cat_data.groupby('Date')['Total pal'].sum()
        
        return {
            'mean': daily_pallets.mean(),
            'std': daily_pallets.std(),
            'median': daily_pallets.median(),
            'min': daily_pallets.min(),
            'max': daily_pallets.max(),
            'total_days': len(daily_pallets)
        }
    
    daily_demand = {
        'FG': {
            'inbound': daily_demand_stats(inbound_year, 'FG'),
            'outbound': daily_demand_stats(outbound_year, 'FG')
        },
        'R&P': {
            'inbound': daily_demand_stats(inbound_year, 'R&P'),
            'outbound': daily_demand_stats(outbound_year, 'R&P')
        }
    }
    
    print(f"\n每日需求统计（托盘 - 全年平均）:")
    for category in ['FG', 'R&P']:
        print(f"\n{category}:")
        inbound_days = daily_demand[category]['inbound']['total_days']
        outbound_days = daily_demand[category]['outbound']['total_days']
        print(f"  入库 - 平均: {daily_demand[category]['inbound']['mean']:.0f}, "
              f"标准差: {daily_demand[category]['inbound']['std']:.0f} (统计{inbound_days}天)")
        print(f"  出库 - 平均: {daily_demand[category]['outbound']['mean']:.0f}, "
              f"标准差: {daily_demand[category]['outbound']['std']:.0f} (统计{outbound_days}天)")
    
    demand_distribution = {
        'hourly_arrival_rate': total_hourly_rate,
        'hourly_arrival_outbound': hourly_arrivals_outbound,
        'hourly_arrival_inbound': hourly_arrivals_inbound,
        'daily_demand': daily_demand
    }
    
    return demand_distribution


def calculate_factory_production_rate(daily_demand):
    """基于每日需求计算工厂24/7生产速率"""
    print("\n" + "=" * 60)
    print("3. 计算工厂生产速率")
    print("=" * 60)
    
    production_rates = {}
    
    for category in ['FG', 'R&P']:
        daily_production = daily_demand[category]['inbound']['mean']
        hourly_rate = daily_production / 24
        
        production_rates[category] = {
            'hourly_rate': hourly_rate,
            'daily_production': daily_production
        }
        
        print(f"\n{category} 工厂生产速率:")
        print(f"  每日生产: {daily_production:.0f} 托盘")
        print(f"  每小时: {hourly_rate:.1f} 托盘/小时（24/7 连续）")
    
    return production_rates


def estimate_buffer_capacity_requirement(production_rate, dc_closed_hours=6):
    """估算缓冲区容量需求（基于DC关闭时段累积量）"""
    print("\n" + "=" * 60)
    print("4. 估算缓冲区容量需求")
    print("=" * 60)
    
    pallets_per_trailer = 33
    buffer_requirements = {}
    
    for category in ['FG', 'R&P']:
        # DC 关闭期间累积的托盘数
        accumulated_pallets = production_rate[category]['hourly_rate'] * dc_closed_hours
        
        # 需要的挂车数（向上取整）
        required_trailers = np.ceil(accumulated_pallets / pallets_per_trailer)
        
        # 加上 20% 安全余量
        recommended_trailers = int(required_trailers * 1.2)
        
        buffer_requirements[category] = {
            'accumulated_pallets': accumulated_pallets,
            'required_trailers': int(required_trailers),
            'recommended_trailers': recommended_trailers
        }
        
        print(f"\n{category} 缓冲区需求 (DC 关闭 {dc_closed_hours} 小时):")
        print(f"  累积托盘: {accumulated_pallets:.0f}")
        print(f"  最低挂车数: {int(required_trailers)}")
        print(f"  推荐挂车数: {recommended_trailers} (含 20% 余量)")
    
    return buffer_requirements


def visualize_hourly_arrival_pattern(hourly_rates):
    """可视化每小时到达模式"""
    print("\n" + "=" * 60)
    print("5. 生成可视化图表")
    print("=" * 60)
    
    fig, ax = plt.subplots(figsize=(12, 6))
    
    hours = sorted(hourly_rates.keys())
    rates = [hourly_rates[h] for h in hours]
    
    ax.bar(hours, rates, color='steelblue', alpha=0.7, edgecolor='navy')
    ax.plot(hours, rates, color='red', marker='o', linewidth=2, markersize=6, label='到达率趋势')
    
    ax.set_xlabel('小时', fontsize=12)
    ax.set_ylabel('平均卡车到达数', fontsize=12)
    ax.set_title('每小时卡车到达分布（基于 2025 年 11 月数据）', fontsize=14, fontweight='bold')
    ax.set_xticks(range(0, 24))
    ax.set_xticklabels([f'{h:02d}:00' for h in range(0, 24)], rotation=45, ha='right')
    ax.grid(axis='y', alpha=0.3)
    ax.legend()
    
    plt.tight_layout()
    fig_path = FIGURES_DIR / 'hourly_arrival_pattern.png'
    plt.savefig(fig_path, dpi=300, bbox_inches='tight')
    print(f"  已保存: {fig_path.name}")
    plt.close()


def extract_dock_capacity_from_timeslot():
    """从48周Timeslot数据提取每小时码头容量"""
    print("\n" + "=" * 60)
    print("提取码头容量数据")
    print("=" * 60)
    
    from openpyxl import load_workbook
    
    # 查找所有Timeslot文件（修正路径：直接在data目录下）
    timeslot_files = list(TIMESLOT_DIR.glob('W*.xlsx'))
    
    if not timeslot_files:
        print(f"警告: 未找到Timeslot文件，路径: {TIMESLOT_DIR / 'W*.xlsx'}")
        return None
    
    print(f"  找到 {len(timeslot_files)} 个周数据文件")
    
    try:
        # 读取并合并所有文件（参照Timeslot.py的逻辑）
        dfs = []
        for file in timeslot_files:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            df = pd.DataFrame(data[1:], columns=data[0])
            
            # 标准化列数
            if len(df.columns) == 31:
                df.insert(7, '0h00', 0)
            if len(df.columns) == 32:
                df = df.iloc[:, :-1]
            
            # 统一列名
            df.columns = [f'col_{i}' for i in range(len(df.columns))]
            dfs.append(df)
        
        merged_df = pd.concat(dfs, ignore_index=True)
        print(f"  合并后总行数: {len(merged_df)}")
        
        # 解析日期
        merged_df['Date_parsed'] = pd.to_datetime(merged_df.iloc[:, 1], errors='coerce', dayfirst=True)
        
        # 修正日期错误（YYYY-M-11 → YYYY-11-M）
        mask = (merged_df['Date_parsed'].dt.year == 2025) & \
               (merged_df['Date_parsed'].dt.day == 11) & \
               (merged_df['Date_parsed'].dt.month <= 10)
        
        if mask.sum() > 0:
            corrected_dates = merged_df.loc[mask, 'Date_parsed'].apply(
                lambda x: pd.Timestamp(year=x.year, month=11, day=x.month)
            )
            merged_df.loc[mask, 'Date_parsed'] = corrected_dates
        
        # 过滤2025年数据
        year_data = merged_df[merged_df['Date_parsed'].dt.year == 2025]
        
        # 过滤有效数据
        year_data = year_data[year_data.iloc[:, 3].notna()]  # col_3: Category不为空
        year_data = year_data[year_data.iloc[:, 5].isin(['Booking taken', 'Available Capacity'])]
        
        print(f"  有效数据行数: {len(year_data)}")
        
        # 时间列索引（col_7到col_30，共24列）
        time_columns = list(range(7, 31))
        
        # 提取码头容量的函数（参照Timeslot.py的extract_timeslot_data）
        def extract_capacity(df, condition_type, category):
            """提取特定条件和类别的容量数据"""
            if condition_type == 'outbound':
                condition_filter = df.iloc[:, 0] == 'Loading'
            else:
                condition_filter = df.iloc[:, 0] == 'Reception'
            
            category_filter = df.iloc[:, 3] == category
            filtered_df = df[condition_filter & category_filter]
            
            # 提取Booking taken和Available Capacity
            booking_taken = filtered_df[filtered_df.iloc[:, 5] == 'Booking taken'].iloc[:, time_columns].sum()
            available_capacity = filtered_df[filtered_df.iloc[:, 5] == 'Available Capacity'].iloc[:, time_columns].sum()
            
            # 总容量 = Booking taken + Available Capacity
            total_capacity = booking_taken + available_capacity
            
            # 计算平均每小时容量（除以天数得到平均值）
            num_days = len(filtered_df[filtered_df.iloc[:, 5] == 'Booking taken']['Date_parsed'].unique())
            if num_days > 0:
                avg_capacity = total_capacity / num_days
            else:
                avg_capacity = total_capacity
            
            return {hour: int(round(avg_capacity.iloc[hour])) if hour < len(avg_capacity) else 0 
                    for hour in range(24)}
        
        # 提取四种组合的容量
        dock_capacity = {
            'FG': {
                'loading': extract_capacity(year_data, 'outbound', 'FG'),
                'reception': extract_capacity(year_data, 'inbound', 'FG')
            },
            'R&P': {
                'loading': extract_capacity(year_data, 'outbound', 'R&P'),
                'reception': extract_capacity(year_data, 'inbound', 'R&P')
            }
        }
        
        # 打印统计信息
        print("成功提取码头容量数据")
        print(f"    FG Loading: 平均 {np.mean(list(dock_capacity['FG']['loading'].values())):.1f} 个码头/小时")
        print(f"    FG Reception: 平均 {np.mean(list(dock_capacity['FG']['reception'].values())):.1f} 个码头/小时")
        print(f"    R&P Loading: 平均 {np.mean(list(dock_capacity['R&P']['loading'].values())):.1f} 个码头/小时")
        print(f"    R&P Reception: 平均 {np.mean(list(dock_capacity['R&P']['reception'].values())):.1f} 个码头/小时")
        
        return dock_capacity
        
    except Exception as e:
        print(f"警告: 提取码头容量失败（异常）: {e}")
        import traceback
        traceback.print_exc()
        return None


def extract_pallet_distribution():
    """从Total Shipments分析托盘数分布"""
    print("\n" + "=" * 60)
    print("分析托盘数分布")
    print("=" * 60)
    
    # 读取 Inbound 和 Outbound 数据（实际sheet名称）
    inbound_df = pd.read_excel(SHIPMENTS_FILE, sheet_name='Inbound Shipments 2025')
    outbound_df = pd.read_excel(SHIPMENTS_FILE, sheet_name='Outbound Shipments 2025')
    
    # 合并所有数据
    all_shipments = pd.concat([inbound_df, outbound_df], ignore_index=True)
    
    print(f"  总批次数: {len(all_shipments)}")
    
    pallet_distribution = {}
    
    for category in ['FG', 'R&P']:
        category_data = all_shipments[all_shipments['Category'] == category]['Total pal'].dropna()
        
        if len(category_data) == 0:
            print(f"警告: {category}: 无数据")
            continue
        
        # 计算分布参数
        pallets = category_data.values
        
        # 基本统计
        mean_val = float(np.mean(pallets))
        std_val = float(np.std(pallets))
        median_val = float(np.median(pallets))
        min_val = float(np.min(pallets))
        max_val = float(np.max(pallets))
        
        # 估算众数（使用核密度估计的峰值）
        try:
            # 使用直方图近似众数
            hist, bin_edges = np.histogram(pallets, bins=50)
            mode_idx = np.argmax(hist)
            mode_val = float((bin_edges[mode_idx] + bin_edges[mode_idx + 1]) / 2)
        except:
            mode_val = median_val
        
        # 确定分布类型
        # 计算偏度 (利用 pandas Series 的 skew 方法，避免引入 scipy)
        skewness = category_data.skew()
        
        # 基于偏度选择分布类型
        if abs(skewness) < 0.5:
            dist_type = 'normal'
        else:
            dist_type = 'triangular'  # 三角分布对偏斜数据更稳健
        
        pallet_distribution[category] = {
            'type': dist_type,
            'min': int(min_val),
            'mode': int(round(mode_val)),
            'max': int(max_val),
            'mean': mean_val,
            'std': std_val,
            'median': median_val,
            'count': len(category_data)
        }
        
        print(f"\n  {category} 托盘数分布:")
        print(f"    样本量: {len(category_data):,} 批次")
        print(f"    分布类型: {dist_type}")
        print(f"    均值: {mean_val:.1f} 托盘")
        print(f"    标准差: {std_val:.1f}")
        print(f"    中位数: {median_val:.1f}")
        print(f"    众数: {mode_val:.0f}")
        print(f"    范围: [{int(min_val)}, {int(max_val)}]")
        print(f"    偏度: {skewness:.2f}")
    
    return pallet_distribution


def extract_monthly_totals_from_kpi():
    """从KPI sheet提取每月总托盘数（权威数据源）
    
    Returns:
        dict: {category: {direction: {month: total_pallets}}}
    """
    print("\n" + "=" * 60)
    print("提取KPI月度总托盘数（权威数据源）")
    print("=" * 60)
    
    df = pd.read_excel(KPI_FILE, sheet_name='Hours & volumes per subgroup', header=None)
    
    monthly_totals = {
        'FG': {'Inbound': {}, 'Outbound': {}},
        'R&P': {'Inbound': {}, 'Outbound': {}}
    }
    
    # 提取每个category的数据
    for category in ['FG', 'R&P']:
        # 查找Inbound和Outbound的行
        for idx, row in df.iterrows():
            if row[0] == f'{category} - pallets - inbound' and row[1] == 'Actual':
                inbound_data = pd.to_numeric(row[2:14], errors='coerce')
                for month_idx, value in enumerate(inbound_data, start=1):
                    if not pd.isna(value):
                        monthly_totals[category]['Inbound'][month_idx] = float(value)
            
            elif row[0] == f'{category} - pallets - outbound' and row[1] == 'Actual':
                outbound_data = pd.to_numeric(row[2:14], errors='coerce')
                for month_idx, value in enumerate(outbound_data, start=1):
                    if not pd.isna(value):
                        monthly_totals[category]['Outbound'][month_idx] = float(value)
    
    # 打印汇总
    print("\nKPI月度总托盘数:")
    for category in ['FG', 'R&P']:
        for direction in ['Inbound', 'Outbound']:
            total = sum(monthly_totals[category][direction].values())
            months = len(monthly_totals[category][direction])
            avg = total / months if months > 0 else 0
            print(f"  {category} {direction}: {months}个月, 年总量={total:,.0f}, 月均={avg:,.0f}")
    
    return monthly_totals


def generate_orders_for_month(month, category, direction, kpi_total_pallet, 
                                shipments_df, pallet_dist, dock_capacity):
    """为指定月份生成校准后的订单数据
    
    Args:
        month: 月份 (1-12)
        category: 'FG' or 'R&P'
        direction: 'Inbound' or 'Outbound'
        kpi_total_pallet: KPI sheet的该月总托盘数
        shipments_df: Total Shipments原始数据
        pallet_dist: 托盘分布参数
        dock_capacity: 码头容量配置
    
    Returns:
        DataFrame: 订单数据
    """
    # 1. 筛选该月、该类别的数据
    shipments_df['Month'] = shipments_df['Date Hour appointement'].dt.month
    month_data = shipments_df[
        (shipments_df['Month'] == month) & 
        (shipments_df['Category'] == category)
    ].copy()
    
    if len(month_data) == 0:
        print(f"    警告: {category} {direction} 月{month} 无shipments数据，跳过")
        return None
    
    # 2. 计算校正比例
    shipments_total = month_data['Total pal'].sum()
    if shipments_total == 0:
        print(f"    警告: {category} {direction} 月{month} shipments总数为0，跳过")
        return None
    
    correction_ratio = kpi_total_pallet / shipments_total
    original_order_count = len(month_data)
    corrected_order_count = int(round(original_order_count * correction_ratio))
    
    print(f"    {category} {direction} 月{month}:")
    print(f"      原始订单数={original_order_count}, 校正后={corrected_order_count}")
    print(f"      校正比例={correction_ratio:.3f}")
    
    # 3. 生成订单
    orders = []
    days_in_month = pd.Period(f'2025-{month:02d}').days_in_month
    orders_per_day = corrected_order_count / days_in_month
    
    # 从pallet分布采样
    dist_params = pallet_dist.get(category, {})
    if dist_params.get('type') == 'triangular':
        sampled_pallets = np.random.triangular(
            dist_params['min'], 
            dist_params['mode'], 
            dist_params['max'], 
            corrected_order_count
        )
    else:  # normal
        sampled_pallets = np.random.normal(
            dist_params['mean'], 
            dist_params['std'], 
            corrected_order_count
        )
    
    # 确保非负且为整数
    sampled_pallets = np.maximum(1, sampled_pallets).astype(int)
    
    # 4. 按比例调整使总和=kpi_total_pallet
    sampled_total = sampled_pallets.sum()
    scale_factor = kpi_total_pallet / sampled_total
    adjusted_pallets = (sampled_pallets * scale_factor).astype(int)
    
    # 最后一个订单补齐差额
    adjusted_pallets[-1] += int(kpi_total_pallet - adjusted_pallets.sum())
    
    # 5. 生成订单记录
    for i in range(corrected_order_count):
        # 均匀分配到每天
        day = int(i / orders_per_day) + 1
        if day > days_in_month:
            day = days_in_month
        
        order = {
            'order_id': f'{category}_{direction}_{month:02d}_{i+1:05d}',
            'month': month,
            'day': day,
            'category': category,
            'direction': direction,
            'pallets': int(adjusted_pallets[i])
        }
        
        # Outbound特有属性
        if direction == 'Outbound':
            # 随机分配region (40-40-20)
            rand = np.random.random()
            if rand < 0.4:
                order['region'] = 'G2_same_day'
                # creation_time: 当日0-12h
                order['creation_hour'] = np.random.uniform(0, 12)
            elif rand < 0.8:
                order['region'] = 'G2_next_day'
                # creation_time: 前一天12-24h
                order['creation_hour'] = np.random.uniform(12, 24) - 24  # 负数表示前一天
            else:
                order['region'] = 'ROW_next_day'
                # creation_time: 前一天0h
                order['creation_hour'] = -24  # 前一天0h
        
        # Inbound特有属性: 从原始数据采样timeslot小时
        if direction == 'Inbound':
            if len(month_data) > 0:
                sample = month_data.sample(1).iloc[0]
                order['timeslot_hour'] = sample['Date Hour appointement'].hour
            else:
                order['timeslot_hour'] = np.random.randint(6, 22)  # 默认6-21h
        
        orders.append(order)
    
    orders_df = pd.DataFrame(orders)
    
    # 6. Outbound: 贪心分配timeslot
    if direction == 'Outbound':
        orders_df = allocate_outbound_timeslots(orders_df, dock_capacity, category)
    
    print(f"      生成订单={len(orders_df)}, 总托盘={orders_df['pallets'].sum():,.0f}")
    
    return orders_df


def allocate_outbound_timeslots(orders_df, dock_capacity, category):
    """贪心算法为Outbound订单分配timeslot
    
    Args:
        orders_df: 订单DataFrame（已有creation_hour和region）
        dock_capacity: 码头容量配置
        category: 'FG' or 'R&P'
    
    Returns:
        orders_df: 添加了timeslot_hour列
    """
    # 计算绝对creation time (以第一天0点为基准)
    orders_df['creation_time_abs'] = orders_df['day'] * 24 + orders_df['creation_hour']
    
    # 按creation_time排序
    orders_df = orders_df.sort_values('creation_time_abs').copy()
    
    # 获取容量配置
    dock_type = 'loading'
    capacity_dict = dock_capacity.get(category, {}).get(dock_type, {})
    
    # 每小时使用计数器 {absolute_hour: count}
    usage_counter = {}
    
    for idx, row in orders_df.iterrows():
        region = row['region']
        creation_abs = row['creation_time_abs']
        day = row['day']
        
        # 确定搜索范围
        if region == 'G2_same_day':
            # creation + 5h 到当日24h
            start_abs = creation_abs + 5
            end_abs = day * 24
        else:  # G2_next_day, ROW_next_day
            # creation后到次日24h
            start_abs = max(creation_abs, day * 24)  # 至少从当天开始
            end_abs = (day + 1) * 24
        
        # 查找最早可用slot
        allocated = False
        for abs_hour in range(int(start_abs), int(end_abs) + 1):
            hour_of_day = abs_hour % 24
            max_capacity = capacity_dict.get(hour_of_day, capacity_dict.get(str(hour_of_day), 0))
            current_usage = usage_counter.get(abs_hour, 0)
            
            if current_usage < max_capacity:
                orders_df.at[idx, 'timeslot_hour'] = hour_of_day
                orders_df.at[idx, 'timeslot_abs'] = abs_hour
                usage_counter[abs_hour] = current_usage + 1
                allocated = True
                break
        
        # 如果无法分配，找最早的可用slot（延误）
        if not allocated:
            for abs_hour in range(int(end_abs) + 1, int(end_abs) + 100):
                hour_of_day = abs_hour % 24
                max_capacity = capacity_dict.get(hour_of_day, capacity_dict.get(str(hour_of_day), 0))
                current_usage = usage_counter.get(abs_hour, 0)
                
                if current_usage < max_capacity:
                    orders_df.at[idx, 'timeslot_hour'] = hour_of_day
                    orders_df.at[idx, 'timeslot_abs'] = abs_hour
                    orders_df.at[idx, 'delayed'] = True
                    usage_counter[abs_hour] = current_usage + 1
                    break
    
    return orders_df


def extract_fte_from_kpi():
    """
    从 KPI sheet 提取人力资源数据
    
    Returns:
        dict: FTE总数和分配
    """
    print("\n" + "=" * 60)
    print("8. 提取人力资源数据")
    print("=" * 60)
    
    df = pd.read_excel(KPI_FILE, sheet_name='Hours & volumes per subgroup', header=None)
    
    # 提取工时数据（修复：读取Actual行本身，而不是下一行Delta）
    def extract_hours(df, category):
        """提取特定类别的工时数据"""
        for idx, row in df.iterrows():
            if row[0] == category and row[1] == 'Actual':
                # 直接读取当前行（Actual行），而不是idx+1（Delta行）
                hours_data = df.iloc[idx, 2:13].values
                hours_data = pd.to_numeric(hours_data, errors='coerce')
                return hours_data[~pd.isna(hours_data)]
        return None
    
    rp_hours = extract_hours(df, 'R&P')
    fg_hours = extract_hours(df, 'FG')
    
    if rp_hours is None or fg_hours is None:
        print("警告: 无法提取工时数据，使用默认值")
        return {
            'fte_total': 125,
            'fte_allocation': {'rp_baseline': 28, 'fg_baseline': 97}
        }
    
    # 计算平均月工时
    rp_avg_hours = np.mean(rp_hours)
    fg_avg_hours = np.mean(fg_hours)
    total_avg_hours = rp_avg_hours + fg_avg_hours
    
    # 标准工时/人/月（假设每人每月176小时，即44小时/周 × 4周）
    hours_per_fte_per_month = 176
    
    # 计算 FTE
    rp_fte = rp_avg_hours / hours_per_fte_per_month
    fg_fte = fg_avg_hours / hours_per_fte_per_month
    total_fte = rp_fte + fg_fte
    
    print(f"\n  平均月工时:")
    print(f"    R&P: {rp_avg_hours:.0f} 小时/月")
    print(f"    FG: {fg_avg_hours:.0f} 小时/月")
    print(f"    总计: {total_avg_hours:.0f} 小时/月")
    
    print(f"\n  计算 FTE (按 {hours_per_fte_per_month} 小时/人/月):")
    print(f"    R&P: {rp_fte:.1f} FTE → {int(round(rp_fte))} 人")
    print(f"    FG: {fg_fte:.1f} FTE → {int(round(fg_fte))} 人")
    print(f"    总计: {total_fte:.1f} FTE → {int(round(total_fte))} 人")
    
    return {
        'fte_total': int(round(total_fte)),
        'fte_allocation': {
            'rp_baseline': int(round(rp_fte)),
            'fg_baseline': int(round(fg_fte))
        }
    }


def generate_simulation_config(efficiency_params, demand_distribution, 
                                production_rates,
                                dock_capacity=None, pallet_distribution=None, 
                                fte_data=None, fte_config=None,
                                monthly_totals=None, orders_file_path=None):
    """生成完整的仿真配置文件"""
    print("\n" + "=" * 60)
    print("生成仿真配置文件")
    print("=" * 60)
    
    # 辅助函数：将numpy类型转换为Python原生类型
    def convert_to_native(obj):
        """递归转换numpy类型为Python原生类型"""
        import numpy as np
        
        if isinstance(obj, dict):
            return {k: convert_to_native(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [convert_to_native(item) for item in obj]
        elif isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return convert_to_native(obj.tolist())
        else:
            return obj
    
    config = {
        'efficiency': {
            'rp_mean': float(efficiency_params['R&P']['mean']),
            'rp_std': float(efficiency_params['R&P']['std']),
            'fg_mean': float(efficiency_params['FG']['mean']),
            'fg_std': float(efficiency_params['FG']['std'])
        },
        'factory_production': {
            'rp_rate': float(production_rates['R&P']['hourly_rate']),
            'fg_rate': float(production_rates['FG']['hourly_rate'])
        },
        # 使用分类别的到达率数据（分Inbound和Outbound）
        'truck_arrival_rates_outbound': convert_to_native(demand_distribution['hourly_arrival_outbound']),
        'truck_arrival_rates_inbound': convert_to_native(demand_distribution['hourly_arrival_inbound']),
        'daily_demand': convert_to_native(demand_distribution['daily_demand'])
    }
    
    # 添加码头容量（如果已提取）
    if dock_capacity is not None:
        config['hourly_dock_capacity'] = convert_to_native(dock_capacity)
        print("已包含码头容量数据")
    else:
        print("警告: 码头容量数据未提取，仿真将使用默认值")
    
    # 添加托盘数分布（如果已提取）
    if pallet_distribution is not None:
        config['pallets_distribution'] = convert_to_native(pallet_distribution)
        print("已包含托盘数分布数据")
    else:
        print("警告: 托盘数分布未提取，仿真将使用默认值")
    
    # 添加FTE配置（优先使用新版本从Input FTE Data.txt）
    if fte_config is not None:
        config['fte_config'] = convert_to_native(fte_config)
        print("已包含FTE配置数据（从Input FTE Data.txt）")
    elif fte_data is not None:
        config['fte_total'] = fte_data['fte_total']
        config['fte_allocation'] = fte_data['fte_allocation']
        print("已包含人力资源数据（旧版本）")
    else:
        print("警告: FTE数据未提取，仿真将使用默认值")
    
    # 添加订单数据路径（新增）
    if orders_file_path is not None:
        config['generated_orders_path'] = str(orders_file_path)
        print(f"已包含订单数据路径: {orders_file_path.name}")
    
    # 添加KPI月度总量（新增）
    if monthly_totals is not None:
        config['kpi_monthly_totals'] = convert_to_native(monthly_totals)
        print("已包含KPI月度总量数据")
    
    # 添加opening hour coefficient默认值（新增）
    config['opening_hour_coefficient'] = 1.0
    print("默认opening_hour_coefficient = 1.0（可在scenario配置中覆盖）")
    
    # 添加数据来源信息
    config['data_source'] = {
        'efficiency_source': 'KPI sheet 2025.xlsx (11 months)',
        'demand_source': 'Total Shipments 2025.xlsx (full year)',
        'dock_capacity_source': 'Timeslot by week W1-W48.xlsx (48 weeks)' if dock_capacity else 'Hardcoded',
        'pallet_distribution_source': 'Total Shipments 2025.xlsx analysis' if pallet_distribution else 'Hardcoded',
        'fte_source': 'KPI sheet 2025.xlsx Hours data' if fte_data else 'Hardcoded',
        'extraction_date': pd.Timestamp.now().strftime('%Y-%m-%d'),
        'year': 2025
    }
    
    # 保存为 JSON 文件
    config_file = OUTPUT_DIR / 'simulation_config.json'
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)
    
    print(f"  已保存: {config_file.name}")
    
    # 保存为 Excel 文件（便于查看）
    params_file = OUTPUT_DIR / 'simulation_parameters.xlsx'
    with pd.ExcelWriter(params_file, engine='openpyxl') as writer:
        # 效率参数
        efficiency_df = pd.DataFrame({
            'Category': ['R&P', 'FG'],
            'Mean Efficiency (pal/hr)': [
                efficiency_params['R&P']['mean'],
                efficiency_params['FG']['mean']
            ],
            'Std Dev': [
                efficiency_params['R&P']['std'],
                efficiency_params['FG']['std']
            ],
            'Min': [
                efficiency_params['R&P']['min'],
                efficiency_params['FG']['min']
            ],
            'Max': [
                efficiency_params['R&P']['max'],
                efficiency_params['FG']['max']
            ]
        })
        efficiency_df.to_excel(writer, sheet_name='Efficiency', index=False)
        
        # 生产速率
        production_df = pd.DataFrame({
            'Category': ['R&P', 'FG'],
            'Hourly Rate (pal/hr)': [
                production_rates['R&P']['hourly_rate'],
                production_rates['FG']['hourly_rate']
            ],
            'Daily Production (pal)': [
                production_rates['R&P']['daily_production'],
                production_rates['FG']['daily_production']
            ]
        })
        production_df.to_excel(writer, sheet_name='Production_Rate', index=False)
        
        # 缓冲区需求（已移除）
        
        # 每小时到达率
        arrival_df = pd.DataFrame({
            'Hour': sorted(demand_distribution['hourly_arrival_rate'].keys()),
            'Arrival Rate (trucks/hr)': [
                demand_distribution['hourly_arrival_rate'][h] 
                for h in sorted(demand_distribution['hourly_arrival_rate'].keys())
            ]
        })
        arrival_df.to_excel(writer, sheet_name='Hourly_Arrivals', index=False)
    
    print(f"  已保存: {params_file.name}")
    
    return config


def main():
    """主函数 - 执行所有数据提取步骤"""
    print("\n" + "="*70)
    print("DC 仿真数据准备脚本 - 完整版（含订单生成）")
    print("从现有数据中提取仿真所需的全部参数 + 生成订单数据")
    print("="*70)
    
    try:
        # 0. 提取FTE配置（从Input FTE Data.txt） - 优先！
        fte_config = extract_fte_from_file()
        
        # 1. 提取效率参数（基于11个月数据）
        efficiency_params = extract_efficiency_parameters()
        
        # 2. 提取需求分布（使用全年数据）
        demand_distribution = extract_demand_distribution(target_year=2025)
        
        # 3. 计算工厂生产速率
        production_rates = calculate_factory_production_rate(demand_distribution['daily_demand'])
        
        # 6. 提取码头容量（48周数据）
        print("\n提取额外数据维度...")
        dock_capacity = None
        try:
            dock_capacity = extract_dock_capacity_from_timeslot()
        except Exception as e:
            print(f"警告: 提取码头容量失败: {e}")
            print("  将使用默认值")
        
        # 7. 分析托盘数分布
        pallet_distribution = None
        try:
            pallet_distribution = extract_pallet_distribution()
        except Exception as e:
            print(f"警告: 分析托盘数分布失败: {e}")
            print("  将使用默认值")
        
        # 8. 提取人力资源数据（旧版本，如果没有Input FTE Data.txt）
        fte_data = None
        if not fte_config:
            try:
                fte_data = extract_fte_from_kpi()
            except Exception as e:
                print(f"警告: 提取人力资源数据失败: {e}")
                print("  将使用默认值")
        
        # ===== 新增：订单生成流程 =====
        print("\n" + "="*70)
        print("开始生成订单数据（新逻辑）")
        print("="*70)
        
        # 9. 提取KPI月度总量
        monthly_totals = None
        try:
            monthly_totals = extract_monthly_totals_from_kpi()
        except Exception as e:
            print(f"错误: 无法提取KPI月度总量: {e}")
            print("  订单生成将被跳过")
        
        # 10. 读取shipments原始数据
        inbound_shipments = pd.read_excel(SHIPMENTS_FILE, sheet_name='Inbound Shipments 2025')
        outbound_shipments = pd.read_excel(SHIPMENTS_FILE, sheet_name='Outbound Shipments 2025')
        inbound_shipments['Date Hour appointement'] = pd.to_datetime(inbound_shipments['Date Hour appointement'])
        outbound_shipments['Date Hour appointement'] = pd.to_datetime(outbound_shipments['Date Hour appointement'])
        
        # 11. 生成所有订单
        all_orders = {}
        if monthly_totals and pallet_distribution and dock_capacity:
            print("\n生成月度订单数据:")
            for category in ['FG', 'R&P']:
                for direction in ['Inbound', 'Outbound']:
                    shipments_df = inbound_shipments if direction == 'Inbound' else outbound_shipments
                    
                    for month in range(1, 13):
                        kpi_total = monthly_totals[category][direction].get(month)
                        if kpi_total is None or kpi_total == 0:
                            continue
                        
                        try:
                            orders_df = generate_orders_for_month(
                                month, category, direction, kpi_total,
                                shipments_df, pallet_distribution, dock_capacity
                            )
                            
                            if orders_df is not None and len(orders_df) > 0:
                                key = f"{category}_{direction}_M{month:02d}"
                                all_orders[key] = orders_df
                        except Exception as e:
                            print(f"    错误: 生成订单失败 - {e}")
            
            # 12. 保存订单数据
            if all_orders:
                orders_output_path = OUTPUT_DIR / 'generated_orders.json'
                
                # 转换为可序列化格式
                orders_serializable = {}
                for key, df in all_orders.items():
                    orders_serializable[key] = df.to_dict(orient='records')
                
                with open(orders_output_path, 'w', encoding='utf-8') as f:
                    json.dump(orders_serializable, f, indent=2, ensure_ascii=False)
                
                print(f"\n✓ 订单数据已保存: {orders_output_path}")
                print(f"  总计生成 {len(all_orders)} 个月度订单文件")
                print(f"  总订单数: {sum(len(df) for df in all_orders.values()):,}")
                print(f"  总托盘数: {sum(df['pallets'].sum() for df in all_orders.values()):,.0f}")
            else:
                orders_output_path = None
                print("\n警告: 未能生成任何订单数据")
        else:
            orders_output_path = None
            print("\n警告: 跳过订单生成（缺少必要数据）")
        
        # ===== 结束订单生成流程 =====
        
        # 13. 生成配置文件（包含所有数据）
        config = generate_simulation_config(
            efficiency_params,
            demand_distribution,
            production_rates,
            dock_capacity=dock_capacity,
            pallet_distribution=pallet_distribution,
            fte_data=fte_data,
            fte_config=fte_config,
            monthly_totals=monthly_totals,
            orders_file_path=orders_output_path if 'orders_output_path' in locals() else None
        )
        
        # 打印汇总
        print("\n" + "="*70)
        print("数据准备完成！")
        print("="*70)
        
        # 统计提取的数据维度
        extracted_params = 4  # 基础4项
        if dock_capacity:
            extracted_params += 1
        if pallet_distribution:
            extracted_params += 1
        if fte_config or fte_data:
            extracted_params += 1
        
        print(f"\n已提取参数类别: {extracted_params}/7")
        print("\n基础参数 (4项):")
        print("  1. 效率参数 (efficiency)")
        print("  2. 工厂生产速率 (factory_production)")
        print("  3. 缓冲区容量 (buffer_capacity)")
        print("  4. 卡车到达率 (truck_arrival_rates)")
        
        if dock_capacity or pallet_distribution or fte_config or fte_data:
            print("\n扩展参数:")
        if dock_capacity:
            print("  5. 码头容量 (hourly_dock_capacity) - 96个值")
        if pallet_distribution:
            print("  6. 托盘数分布 (pallets_distribution) - 12个值")
        if fte_config:
            print("  7. FTE配置 (fte_config) - 从Input FTE Data.txt")
        elif fte_data:
            print("  7. 人力资源 (fte_total, fte_allocation) - 3个值")
        
        if not all([dock_capacity, pallet_distribution, (fte_config or fte_data)]):
            print("\n缺失参数:")
            if not dock_capacity:
                print("  - 码头容量 (仿真将使用硬编码估计值)")
            if not pallet_distribution:
                print("  - 托盘数分布 (仿真将使用硬编码估计值)")
            if not (fte_config or fte_data):
                print("  - FTE配置 (仿真将使用硬编码估计值)")
        
        print("\n生成的文件:")
        print(f"  1. {OUTPUT_DIR.relative_to(PROJECT_ROOT) / 'simulation_config.json'} - 仿真配置（JSON 格式）")
        print(f"  2. {OUTPUT_DIR.relative_to(PROJECT_ROOT) / 'simulation_parameters.xlsx'} - 参数汇总表（Excel）")
        print("\n下一步:")
        print("  运行 dc_simulation.py 进行仿真分析")
        print("="*70 + "\n")
        
        return config
        
    except FileNotFoundError as e:
        print(f"\n错误: 找不到数据文件 - {e}")
        print("请确保以下文件存在:")
        print(f"  - {KPI_FILE}")
        print(f"  - {SHIPMENTS_FILE}")
        print(f"  - {DATA_DIR / 'Timeslot by week'} (包含W1-W48.xlsx)")
        return None
    except Exception as e:
        print(f"\n发生错误: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == '__main__':
    config = main()
